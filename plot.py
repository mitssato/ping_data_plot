import main
import data_handle
import sys
import openpyxl as px
from openpyxl import Workbook, load_workbook
from openpyxl.chart import ScatterChart, LineChart, Reference, Series
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties, Font, RichTextProperties
from openpyxl.drawing.line import LineProperties
from copy import deepcopy

excel_path = main.excel_output_path
wb = px.load_workbook(excel_path)
summary_sheet = wb["summary"]

def plot_sheet_fun(title,col,row):
    plot_sheet = wb.create_sheet(title)
    values = Reference(summary_sheet, min_col = col, min_row=2, max_col=col,max_row=row)
    chart = LineChart()
    chart.add_data(values)
    plot_sheet.add_chart(chart, "A1")

plot_sheet_fun("bandwidth",3,data_handle.line_num)
plot_sheet_fun("jitter",4,data_handle.line_num)
plot_sheet_fun("loss",5,data_handle.line_num)


wb.save(excel_path)
wb.close()