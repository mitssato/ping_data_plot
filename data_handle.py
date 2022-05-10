import sys
from time import time
import openpyxl as px
import main

#ファイルを開く
path = main.text_input_path
excel_path = main.excel_output_path
file = open(path)

#すべての行をリストとして読み込み
lines = file.readlines()

#ファイルを閉じる
file.close()

#ブックを新規作成。Workbookの'w'は必ず大文字。
wb = px.Workbook()

#シートを取得
data_sheet = wb.create_sheet(title='raw_data')
sum_sheet = wb.create_sheet(title='summary')
worksheet = wb.remove(wb['Sheet'])

# 変数の定義
num = 0
data_space = 3
sum_space = main.interval + data_space

sum_sheet.cell(1,2,"Time:")
sum_sheet.cell(1,3,"Bandwidth(Kbits/sec):")

for i in range(len(lines)): # 得られたデータの行数分、実行
    lines[i] = lines[i].replace("Interval", "interval -") # dataの列を整理
    list = lines[i].split()
    if len(list) == 0: 
        list = "-"
    print(list)
    if list[0] == '1':
        temp_list = lines[i+sum_space].split()
        # print rtt ----
        rtt = temp_list[6]
        sum_sheet.cell(num+2,3,float(rtt))
        # ---- print rtt end
        num = num + 1 # num per test
    for j in range(len(list)):
        data_sheet.cell(i+1,j+1,value=list[j]) # raw dataの出力

line_num = num + 1
wb.save(excel_path)
wb.close()